"""
智能铺货 - Excel 解析 -> 铺货 JSON

用 openpyxl 读取单元格，把每个可见 sheet 当作一个模板来生成 JSON。
全程固定规则，不调用 AI，不依赖历史 template_rules.json。Chatbot 只解析替换内容，不决定最终替换动作。
SKU / 货号等图片字段使用斜体，或“类型”列填写“海报”时，表示该行手动处理，不进入自动替换 JSON。
“类型”列可按分类自动给图片字段追加素材后缀，如“模特图”追加 -M、“PNG”追加 -P。

模式规则：
  - 全表没有黄色标记 -> full：每个 sheet 作为一个模板 job，输出全量 values
  - 任意 sheet 存在黄色标记 -> patch：只输出标黄字段，但保留 sheet/template/module 上下文
"""
from __future__ import annotations

import io
import re
import time
from typing import Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter

from .models import SmartDistributeResponse

_YELLOW_HEXES = {"FFFF00", "FFF2CC", "FFEB9C", "FFE066", "FFD700"}
_RED_HEXES = {"FF0000", "C00000", "E06666", "F4CCCC", "EA9999"}
_MARK_LABELS = {"yellow": "黄色增量", "red": "红色增量"}
_MODULE_COL_KEYWORDS = ["模块", "分类", "品类", "分组"]
_IMAGE_FIELD_KEYWORDS = ["sku", "SKU", "货号", "商品编码", "款号", "图片", "素材"]
_EXCLUDED_FIELD_NAMES = {"序号", "库存", "是否下架", "是否上架", "主推款", "SPU", "spu", "吊牌价", "划线价", "零售价"}
_SUBMODULE_HEADER_KEYWORDS = {"序号", "子模块", "二级模块", "二级分类"}
_TYPE_HEADER_NAMES = {"类型"}
_SKIP_TYPE_VALUES = {"海报", "海报图", "海报款"}
_TYPE_SUFFIX_MAP = {
    "模特": "-M",
    "模特图": "-M",
    "人模": "-M",
    "人物": "-M",
    "人像": "-M",
    "PNG": "-P",
    "PNG带阴影": "-S",
    "PNG 带阴影": "-S",
    "带阴影": "-S",
    "阴影": "-S",
    "白底": "-W",
    "白底图": "-W",
    "白图": "-W",
    "一双鞋": "-X2",
}
_TYPE_SUFFIX_LABELS = {
    "-M": "模特",
    "-P": "PNG",
    "-S": "PNG 带阴影",
    "-W": "白底",
    "-X2": "一双鞋",
}


def _norm(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).lower()


_SKIP_TYPE_KEYS = {_norm(v) for v in _SKIP_TYPE_VALUES}
_TYPE_SUFFIX_KEYS = {_norm(k): v for k, v in _TYPE_SUFFIX_MAP.items()}


def _dimension_hidden(dim) -> bool:
    if dim is None:
        return False
    if getattr(dim, "hidden", False):
        return True
    size = getattr(dim, "width", None)
    if size is None:
        size = getattr(dim, "height", None)
    return size is not None and float(size) <= 0.1


def _is_row_hidden(ws, row: int) -> bool:
    return _dimension_hidden(ws.row_dimensions.get(row))


def _is_col_hidden(ws, col: int) -> bool:
    letter = get_column_letter(col)
    if _dimension_hidden(ws.column_dimensions.get(letter)):
        return True
    for dim in ws.column_dimensions.values():
        min_col = getattr(dim, "min", None)
        max_col = getattr(dim, "max", None)
        if min_col and max_col and min_col <= col <= max_col and _dimension_hidden(dim):
            return True
    return False


def _visible_cell_value(ws, row: int, col: int) -> str:
    if _is_row_hidden(ws, row) or _is_col_hidden(ws, col):
        return ""
    merged = _get_merged_range(ws, row, col)
    if merged is not None:
        return merged
    return _cell_str(ws.cell(row=row, column=col))


def _color_hex(cell) -> Optional[str]:
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return None
    fg = fill.fgColor
    if fg is None:
        return None
    if fg.type == "rgb" and fg.rgb:
        raw = str(fg.rgb).upper()
        if raw in ("0", "00000000", "FFFFFF", "FFFFFFFF"):
            return None
        return raw[-6:]
    if fg.type == "indexed" and fg.indexed is not None:
        indexed = int(fg.indexed)
        colors = getattr(openpyxl.styles.colors, "COLOR_INDEX", ())
        if 0 <= indexed < len(colors):
            raw = str(colors[indexed] or "").upper()
            return raw[-6:] if raw else None
    if fg.type == "theme" and fg.tint is not None:
        return None
    return None


def _detect_mark_color(cell) -> Optional[str]:
    hex_val = _color_hex(cell)
    if not hex_val or len(hex_val) != 6:
        return None
    if hex_val in _YELLOW_HEXES:
        return "yellow"
    if hex_val in _RED_HEXES:
        return "red"
    try:
        r = int(hex_val[0:2], 16)
        g = int(hex_val[2:4], 16)
        b = int(hex_val[4:6], 16)
    except ValueError:
        return None
    if r >= 210 and g >= 170 and b <= 120:
        return "yellow"
    if r >= 180 and g <= 140 and b <= 140 and r >= g + 50:
        return "red"
    return None


def _get_merged_range(ws, row: int, col: int) -> Optional[str]:
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            if _is_row_hidden(ws, row) or _is_col_hidden(ws, col):
                return None
            if _is_row_hidden(ws, rng.min_row) or _is_col_hidden(ws, rng.min_col):
                return None
            cell = ws.cell(row=rng.min_row, column=rng.min_col)
            return str(cell.value or "").strip() if cell.value is not None else None
    return None


def _cell_str(cell) -> str:
    v = cell.value
    if v is None:
        return ""
    return str(v).strip() if isinstance(v, str) else str(v).strip()


def _is_na(value: str) -> bool:
    return str(value or "").strip().upper() in {"#N/A", "#REF!", "#VALUE!", "#DIV/0!", "N/A", "NULL"}


def _is_module_col(header: str) -> bool:
    cleaned = _norm(header)
    return any(_norm(kw) == cleaned for kw in _MODULE_COL_KEYWORDS)


def _is_numeric_like(value: str) -> bool:
    return bool(re.fullmatch(r"[\d\s.,，、\-—_]+", str(value or "").strip()))


def _safe_name(value: str) -> str:
    return re.sub(r"[\\/:\*\?\"<>\|]+", "_", str(value or "").strip()) or "默认"


def _with_type_suffix(value: str, suffix: str) -> str:
    text = str(value or "").strip()
    if not text:
        return text
    suffix_key = _norm(suffix)
    if _norm(text).endswith(suffix_key) or _norm(text).endswith(suffix_key.replace("-", "__", 1)):
        return text
    return f"{text}{suffix}"


class SmartDistributor:
    """智能铺货核心处理类。"""

    def process(self, file_bytes: bytes, filename: str, mode: str = "full") -> SmartDistributeResponse:
        mode = "patch" if str(mode or "full").lower() == "patch" else "full"
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, keep_links=False)
        warnings: list[str] = []
        sheets: list[dict] = []
        visible_count = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.sheet_state in ("hidden", "veryHidden"):
                continue
            visible_count += 1
            parsed = self._parse_sheet(ws, sheet_name, warnings)
            if parsed:
                sheets.append(parsed)

        if visible_count == 0:
            warnings.append("未检测到可处理的可见 Sheet，请检查表格是否隐藏了工作表。")

        jobs: list[dict] = []
        if mode == "patch":
            for sheet in sheets:
                for mark_type in ("yellow", "red"):
                    job = self._build_sheet_job(sheet, patch_mode=True, mark_type=mark_type)
                    if job and job.get("modules"):
                        jobs.append(job)
            if not jobs:
                warnings.append("增量模式未检测到黄色或红色标记，请检查表格标记。")
        else:
            for sheet in sheets:
                job = self._build_sheet_job(sheet, patch_mode=False)
                if job and job.get("modules"):
                    jobs.append(job)

        if not jobs:
            warnings.append("未生成任何铺货任务，请检查表格内容")
        summary = self._build_response_summary(jobs)

        return SmartDistributeResponse(
            schemaVersion="1.0",
            mode=mode,
            source={
                "fileName": filename,
                "generatedAt": time.strftime("%Y-%m-%dT%H:%M:%S+08:00", time.localtime()),
                "generator": "internal-chatbot",
            },
            defaults={
                "layerOrder": "panel",
                "savePolicy": "overwrite",
                "exportMode": "moduleGroup",
                "exportFormat": "png",
            },
            summary=summary,
            jobs=jobs,
            warnings=warnings,
        )

    def _parse_sheet(self, ws, sheet_name: str, warnings: list[str]) -> Optional[dict]:
        header_row = self._find_header_row(ws)
        if header_row is None:
            warnings.append(f"Sheet[{sheet_name}] 未检测到有效表头，已跳过")
            return None

        headers, module_col_idx, excluded_cols, type_col_idx = self._scan_headers(ws, header_row)
        if not headers:
            warnings.append(f"Sheet[{sheet_name}] 没有找到有效数据列，已跳过")
            return None

        field_order = [headers[col] for col in sorted(headers)]
        image_field = self._detect_image_field(field_order)
        last_row = self._find_last_data_row(ws, header_row, headers)
        submodule_col_idx = self._find_submodule_col(ws, header_row, last_row, excluded_cols)
        modules = self._detect_modules(ws, header_row, module_col_idx, submodule_col_idx, last_row)
        rows_data, marked_cells, skipped_rows, skipped_type_counts = self._read_rows(
            ws, header_row, headers, image_field, last_row, type_col_idx
        )
        if not rows_data:
            warnings.append(f"Sheet[{sheet_name}] 没有有效数据行")
            return None

        module_groups = self._group_by_module(rows_data, modules, headers, module_col_idx, submodule_col_idx)
        return {
            "sheetName": sheet_name,
            "templateName": sheet_name,
            "fieldOrder": field_order,
            "imageField": image_field,
            "modules": module_groups,
            "markedCells": marked_cells,
            "skippedRows": skipped_rows,
            "skippedTypeCounts": skipped_type_counts,
        }

    def _find_header_row(self, ws) -> Optional[int]:
        candidates: list[tuple[int, int]] = []
        for row_idx in range(1, min(ws.max_row + 1, 21)):
            visible_values = []
            module_col = None
            for col in range(1, ws.max_column + 1):
                value = _visible_cell_value(ws, row_idx, col)
                if not value:
                    continue
                visible_values.append(value)
                if _is_module_col(value):
                    module_col = col
            if module_col is not None:
                return row_idx
            if any(any(_norm(keyword) in _norm(value) for keyword in _IMAGE_FIELD_KEYWORDS) for value in visible_values):
                candidates.append((row_idx, len(visible_values)))
        for row_idx, count in candidates:
            if count >= 2:
                return row_idx
        for row_idx in range(1, min(ws.max_row + 1, 21)):
            count = sum(1 for col in range(1, ws.max_column + 1) if _visible_cell_value(ws, row_idx, col))
            if count >= 2:
                return row_idx
        return None

    def _scan_headers(self, ws, header_row: int) -> tuple[dict[int, str], Optional[int], dict[int, str], Optional[int]]:
        headers: dict[int, str] = {}
        excluded_cols: dict[int, str] = {}
        module_col_idx = None
        type_col_idx = None
        for col_idx in range(1, ws.max_column + 1):
            raw = _visible_cell_value(ws, header_row, col_idx)
            if raw and _is_module_col(raw):
                module_col_idx = col_idx
                break
        start_col = module_col_idx or 1
        for col_idx in range(start_col, ws.max_column + 1):
            raw = _visible_cell_value(ws, header_row, col_idx)
            if not raw:
                continue
            if _is_module_col(raw):
                module_col_idx = col_idx
            elif raw in _TYPE_HEADER_NAMES:
                type_col_idx = col_idx
            elif raw in _EXCLUDED_FIELD_NAMES:
                excluded_cols[col_idx] = raw
            else:
                headers[col_idx] = raw
        return headers, module_col_idx, excluded_cols, type_col_idx

    def _detect_image_field(self, field_order: list[str]) -> str:
        for field in field_order:
            nf = _norm(field)
            if any(_norm(keyword) == nf for keyword in _IMAGE_FIELD_KEYWORDS):
                return field
        for field in field_order:
            nf = _norm(field)
            if any(_norm(keyword) in nf for keyword in _IMAGE_FIELD_KEYWORDS):
                return field
        return field_order[0]

    def _find_last_data_row(self, ws, header_row: int, headers: dict[int, str]) -> int:
        blank_run = 0
        last_row = header_row
        for row_idx in range(header_row + 1, ws.max_row + 1):
            if _is_row_hidden(ws, row_idx):
                continue
            has_any = any(_visible_cell_value(ws, row_idx, col_idx) for col_idx in headers)
            if has_any:
                blank_run = 0
                last_row = row_idx
            else:
                blank_run += 1
                if blank_run >= 15:
                    break
        return last_row

    def _read_rows(self, ws, header_row: int, headers: dict[int, str], image_field: str,
                   last_row: int, type_col_idx: Optional[int]) -> tuple[list[dict], dict[str, set], set, dict[str, int]]:
        rows_data = []
        marked_cells = {"yellow": set(), "red": set()}
        skipped_rows = set()
        skipped_type_counts: dict[str, int] = {}
        image_col = next((col for col, field in headers.items() if field == image_field), min(headers.keys()))
        for row_idx in range(header_row + 1, last_row + 1):
            if _is_row_hidden(ws, row_idx):
                continue
            type_value = ""
            if type_col_idx is not None:
                type_value = _visible_cell_value(ws, row_idx, type_col_idx)
            type_key = _norm(type_value)
            skip_image = False
            if type_key in _SKIP_TYPE_KEYS:
                skip_image = True
                skipped_rows.add(row_idx)
                label = type_value or "跳过"
                skipped_type_counts[label] = skipped_type_counts.get(label, 0) + 1
            suffix = None if skip_image else _TYPE_SUFFIX_KEYS.get(type_key)
            suffix_label = _TYPE_SUFFIX_LABELS.get(suffix or "", type_value)
            image_cell = ws.cell(row=row_idx, column=image_col)
            image_value = _visible_cell_value(ws, row_idx, image_col)
            if image_value and image_cell.font and image_cell.font.italic:
                skip_image = True
                skipped_rows.add(row_idx)

            row_values = {"_row": row_idx}
            if skip_image:
                row_values["_skipImage"] = True
            if suffix:
                row_values["_typeValue"] = type_value
                row_values["_typeSuffix"] = suffix
                row_values["_typeLabel"] = suffix_label
            has_any = False
            has_error = False
            for col_idx, field_name in headers.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                val = _visible_cell_value(ws, row_idx, col_idx)
                mark = _detect_mark_color(cell)
                if mark:
                    marked_cells[mark].add((row_idx, field_name))
                if field_name == image_field:
                    if skip_image:
                        val = ""
                        image_value = ""
                    elif suffix:
                        val = _with_type_suffix(val, suffix)
                        image_value = val
                row_values[field_name] = val
                has_any = has_any or bool(val)
                if not (skip_image and field_name == image_field):
                    has_error = has_error or _is_na(val)
            if not has_any or (not skip_image and not image_value) or has_error:
                continue
            rows_data.append(row_values)
        return rows_data, marked_cells, skipped_rows, skipped_type_counts

    def _find_submodule_col(self, ws, header_row: int, last_row: int, excluded_cols: dict[int, str]) -> Optional[int]:
        """在被排除列里识别“二级模块”列，避免把它误当替换字段。"""
        for col_idx, header in excluded_cols.items():
            if header not in _SUBMODULE_HEADER_KEYWORDS:
                continue
            for rng in ws.merged_cells.ranges:
                if not (rng.min_col <= col_idx <= rng.max_col):
                    continue
                if rng.min_row <= header_row or rng.max_row < header_row + 1 or rng.min_row > last_row:
                    continue
                if rng.max_row - rng.min_row + 1 < 2:
                    continue
                value = _visible_cell_value(ws, rng.min_row, rng.min_col)
                if value and not _is_numeric_like(value):
                    return col_idx
        return None

    def _detect_modules(self, ws, header_row: int, module_col_idx: Optional[int],
                        submodule_col_idx: Optional[int], last_row: int) -> list[dict]:
        if module_col_idx is None and submodule_col_idx is None:
            return []
        modules = []
        current_key = None
        current_payload = None
        start = None
        for row_idx in range(header_row + 1, last_row + 1):
            if _is_row_hidden(ws, row_idx):
                continue
            parent = ""
            if module_col_idx is not None:
                parent = _visible_cell_value(ws, row_idx, module_col_idx)
            submodule = ""
            if submodule_col_idx is not None:
                submodule = _visible_cell_value(ws, row_idx, submodule_col_idx)
            if not parent and not submodule:
                continue
            module_name = "/".join(part for part in (parent, submodule) if part) or "默认"
            key = (parent, submodule)
            payload = {"moduleName": module_name, "parentModule": parent, "subModule": submodule}
            if current_key is None:
                current_key, current_payload, start = key, payload, row_idx
            elif key != current_key:
                modules.append({**current_payload, "rowStart": start, "rowEnd": row_idx - 1})
                current_key, current_payload, start = key, payload, row_idx
        if current_payload is not None and start is not None:
            modules.append({**current_payload, "rowStart": start, "rowEnd": last_row})
        return modules

    def _group_by_module(self, rows_data: list[dict], modules: list[dict], headers: dict[int, str],
                         module_col_idx: Optional[int], submodule_col_idx: Optional[int]) -> list[dict]:
        if not modules:
            row_start = min(r["_row"] for r in rows_data)
            row_end = max(r["_row"] for r in rows_data)
            return [{
                "moduleName": "默认",
                "rowStart": row_start,
                "rowEnd": row_end,
                "excelRange": self._excel_range(row_start, row_end, headers, module_col_idx, submodule_col_idx),
                "rowCount": len(rows_data),
                "rows": rows_data,
            }]
        result = []
        for mod in modules:
            mr = [r for r in rows_data if mod["rowStart"] <= r["_row"] <= mod["rowEnd"]]
            if mr:
                row_start = min(r["_row"] for r in mr)
                row_end = max(r["_row"] for r in mr)
                result.append({
                    "moduleName": mod["moduleName"],
                    "parentModule": mod.get("parentModule") or "",
                    "subModule": mod.get("subModule") or "",
                    "rowStart": row_start,
                    "rowEnd": row_end,
                    "excelRange": self._excel_range(row_start, row_end, headers, module_col_idx, submodule_col_idx),
                    "rowCount": len(mr),
                    "rows": mr,
                })
        return result

    def _excel_range(self, row_start: int, row_end: int, headers: dict[int, str],
                     module_col_idx: Optional[int], submodule_col_idx: Optional[int]) -> str:
        cols = list(headers.keys())
        if module_col_idx:
            cols.append(module_col_idx)
        if submodule_col_idx:
            cols.append(submodule_col_idx)
        if not cols:
            return ""
        return f"{get_column_letter(min(cols))}{row_start}:{get_column_letter(max(cols))}{row_end}"

    def _summarize_rows(self, rows: list[dict]) -> dict:
        suffix_counts: dict[str, int] = {}
        type_counts: dict[str, int] = {}
        for row in rows:
            suffix = row.get("_typeSuffix")
            label = row.get("_typeLabel")
            if suffix:
                suffix_counts[suffix] = suffix_counts.get(suffix, 0) + 1
            if label:
                type_counts[label] = type_counts.get(label, 0) + 1
        return {
            "skuCount": len(rows),
            "specialMarkedCount": sum(type_counts.values()),
            "typeCounts": type_counts,
            "suffixCounts": suffix_counts,
        }

    def _build_response_summary(self, jobs: list[dict]) -> dict:
        summary = {
            "templateCount": len(jobs),
            "moduleCount": 0,
            "skuCount": 0,
            "fieldCount": 0,
            "totalSlots": 0,
            "skippedRows": 0,
            "specialMarkedCount": 0,
            "typeCounts": {},
            "suffixCounts": {},
        }
        for job in jobs:
            js = job.get("summary") or {}
            summary["moduleCount"] += js.get("moduleCount", 0)
            summary["skuCount"] += js.get("skuCount", 0)
            summary["fieldCount"] = max(summary["fieldCount"], js.get("fieldCount", 0))
            summary["totalSlots"] += js.get("totalSlots", 0)
            summary["skippedRows"] += js.get("skippedRows", 0)
            summary["specialMarkedCount"] += js.get("specialMarkedCount", 0)
            for key, value in (js.get("typeCounts") or {}).items():
                summary["typeCounts"][key] = summary["typeCounts"].get(key, 0) + value
            for key, value in (js.get("suffixCounts") or {}).items():
                summary["suffixCounts"][key] = summary["suffixCounts"].get(key, 0) + value
        return summary

    def _build_sheet_job(self, sheet: dict, patch_mode: bool, mark_type: Optional[str] = None) -> Optional[dict]:
        modules = []
        sheet_type_counts: dict[str, int] = {}
        sheet_suffix_counts: dict[str, int] = {}
        sheet_sku_count = 0
        sheet_slot_count = 0
        for mod in sheet["modules"]:
            entries = self._build_entries(sheet, mod, patch_mode, mark_type)
            if patch_mode and not entries:
                continue
            row_summary = self._summarize_rows(mod["rows"])
            module_summary = {
                **row_summary,
                "fieldCount": len(sheet["fieldOrder"]),
                "totalSlots": len(entries),
                "skippedRows": 0,
            }
            module_payload = {
                "moduleName": mod["moduleName"],
                "targetGroup": mod["moduleName"],
                "excelRange": mod["excelRange"],
                "rowCount": mod["rowCount"],
                "expectedLayerCount": len(entries),
                "exportName": f"{sheet['sheetName']}_{_safe_name(mod['moduleName'])}",
                "summary": module_summary,
            }
            if mod.get("parentModule"):
                module_payload["parentModule"] = mod["parentModule"]
            if mod.get("subModule"):
                module_payload["subModule"] = mod["subModule"]
            module_payload["patches" if patch_mode else "values"] = entries
            modules.append(module_payload)
            sheet_sku_count += row_summary["skuCount"]
            sheet_slot_count += len(entries)
            for key, value in row_summary["typeCounts"].items():
                sheet_type_counts[key] = sheet_type_counts.get(key, 0) + value
            for key, value in row_summary["suffixCounts"].items():
                sheet_suffix_counts[key] = sheet_suffix_counts.get(key, 0) + value
        if not modules:
            return None
        for key, value in (sheet.get("skippedTypeCounts") or {}).items():
            sheet_type_counts[key] = sheet_type_counts.get(key, 0) + value
        skipped_rows = sorted(sheet.get("skippedRows") or [])
        sheet_summary = {
            "templateName": sheet["templateName"],
            "moduleCount": len(modules),
            "skuCount": sheet_sku_count,
            "fieldCount": len(sheet["fieldOrder"]),
            "totalSlots": sheet_slot_count,
            "expectedLayerCount": sheet_slot_count,
            "skippedRows": len(skipped_rows),
            "specialMarkedCount": sum(sheet_suffix_counts.values()),
            "typeCounts": sheet_type_counts,
            "suffixCounts": sheet_suffix_counts,
        }
        job = {
            "sheetName": sheet["sheetName"],
            "templateName": sheet["templateName"],
            "fieldOrder": sheet["fieldOrder"],
            "summary": sheet_summary,
            "skippedRows": skipped_rows,
            "modules": modules,
        }
        if patch_mode and mark_type:
            job["batchType"] = mark_type
            job["batchLabel"] = _MARK_LABELS.get(mark_type, mark_type)
        return job

    def _build_entries(self, sheet: dict, mod: dict, patch_mode: bool, mark_type: Optional[str] = None) -> list[dict]:
        items = []
        field_order = sheet["fieldOrder"]
        image_field = sheet["imageField"]
        marked_cells = (sheet.get("markedCells") or {}).get(mark_type or "", set())
        for ri, row_data in enumerate(mod["rows"]):
            skip_image = bool(row_data.get("_skipImage"))
            for field_name in field_order:
                row_idx = row_data["_row"]
                if skip_image and field_name == image_field:
                    continue
                if patch_mode and (row_idx, field_name) not in marked_cells:
                    continue
                entry = {
                    "slotIndex": len(items) + 1,
                    "field": field_name,
                    "type": "image" if field_name == image_field else "text",
                    "value": row_data.get(field_name, ""),
                    "rowIndexInModule": ri + 1,
                    "sourceRow": row_idx,
                    "changed": patch_mode,
                }
                if patch_mode and mark_type:
                    entry["markType"] = mark_type
                items.append(entry)
        return items
